#!/usr/bin/env python3
"""
Sync Yahoo Fantasy data to data.js for the Warrior Poets site.

Usage:
    python3 sync_yahoo_data.py 2024          # Sync 2024 season data
    python3 sync_yahoo_data.py 2024 --dry-run  # Preview without writing
"""

import sys
import json
import re
from pathlib import Path

import openpyxl

from yahoo_fantasy import YahooFantasyAPI, extract_weekly_points

DATA_JS_PATH = Path(__file__).parent / "js" / "data.js"
BACKUP_XLSX_PATH = Path(__file__).parent / "WarriorPoets_Backup.xlsx"

# Points table layout in the backup workbook: row 6-19 is the 14-team weekly
# points grid, column E (5) is week 1, column U (21) is week 17.
POINTS_TABLE_FIRST_ROW = 6
POINTS_TABLE_LAST_ROW = 19
POINTS_TABLE_FIRST_WEEK_COL = 5


def read_data_js():
    """Read and parse the existing data.js file."""
    content = DATA_JS_PATH.read_text()

    # Extract the JSON object from the JS file
    # Remove "const LEAGUE_DATA = " prefix and trailing semicolon
    json_str = content.replace("const LEAGUE_DATA = ", "").rstrip().rstrip(";")

    return json.loads(json_str)


def write_data_js(data):
    """Write data back to data.js file."""
    json_str = json.dumps(data, indent=2)
    content = f"const LEAGUE_DATA = {json_str};\n"
    DATA_JS_PATH.write_text(content)


def write_backup_xlsx(year, weekly_points, dry_run=False):
    """Write weekly points into the matching season tab of the backup xlsx.

    The tab's formulas (totals, ranks, sidebets) recalculate on their own
    from these raw point values, so only the weekly point cells are touched.
    """
    if not BACKUP_XLSX_PATH.exists():
        print(f"\n[backup xlsx] {BACKUP_XLSX_PATH} not found, skipping.")
        return

    wb = openpyxl.load_workbook(BACKUP_XLSX_PATH, data_only=False)
    if year not in wb.sheetnames:
        print(f"\n[backup xlsx] No '{year}' tab in {BACKUP_XLSX_PATH.name}, skipping.")
        return

    ws = wb[year]

    owner_row = {}
    for row in range(POINTS_TABLE_FIRST_ROW, POINTS_TABLE_LAST_ROW + 1):
        owner = ws.cell(row=row, column=4).value  # column D
        if owner:
            owner_row[owner] = row

    print(f"\n--- Backup xlsx changes for {year} ---")
    changed = False
    for member, weeks in weekly_points.items():
        row = owner_row.get(member)
        if row is None:
            print(f"  Skipping {member}: no matching row in the '{year}' tab")
            continue
        for week_str, points in weeks.items():
            week_num = int(week_str)
            if week_num < 1 or week_num > 17:
                continue
            col = POINTS_TABLE_FIRST_WEEK_COL + week_num - 1
            cell = ws.cell(row=row, column=col)
            if cell.value != points:
                print(f"  {member} Week {week_str}: {cell.value} → {points}")
                changed = True
                if not dry_run:
                    cell.value = points

    if not changed:
        print("  No changes.")

    if dry_run:
        print("[DRY RUN] No changes written.")
        return

    if changed:
        wb.save(BACKUP_XLSX_PATH)
        print(f"Saved {BACKUP_XLSX_PATH}")


def sync_season(year, dry_run=False):
    """Sync data for a specific season from Yahoo to data.js."""
    print(f"Syncing {year} season data from Yahoo Fantasy...")

    # Authenticate and fetch data
    api = YahooFantasyAPI()
    if not api.authenticate():
        print("Authentication failed!")
        return False

    # Get weekly points from Yahoo
    weekly_points, team_map = extract_weekly_points(api, year)

    print(f"\nFetched data for {len(weekly_points)} members")

    # Read existing data.js
    print(f"\nReading {DATA_JS_PATH}...")
    data = read_data_js()

    # Check if season exists
    if year not in data.get("seasons", {}):
        print(f"Warning: Season {year} not found in data.js. Creating new entry.")
        data["seasons"][year] = {}

    # Update weekly points
    old_points = data["seasons"][year].get("weeklyPoints", {})
    data["seasons"][year]["weeklyPoints"] = weekly_points

    # Show diff
    print(f"\n--- Changes for {year} ---")
    for member, weeks in weekly_points.items():
        old_member_points = old_points.get(member, {})
        for week, points in weeks.items():
            old_val = old_member_points.get(week)
            if old_val != points:
                if old_val is None:
                    print(f"  {member} Week {week}: NEW → {points}")
                else:
                    print(f"  {member} Week {week}: {old_val} → {points}")

    write_backup_xlsx(year, weekly_points, dry_run)

    if dry_run:
        print("\n[DRY RUN] No changes written.")
        return True

    # Write back
    print(f"\nWriting to {DATA_JS_PATH}...")
    write_data_js(data)
    print("Done!")

    return True


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 sync_yahoo_data.py <year> [--dry-run]")
        print("Example: python3 sync_yahoo_data.py 2024")
        sys.exit(1)

    year = sys.argv[1]
    dry_run = "--dry-run" in sys.argv

    success = sync_season(year, dry_run)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
